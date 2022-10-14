#!/usr/bin/python3
from bisect import bisect_right
from openpyxl import load_workbook

score=[]
wb = load_workbook(filename="student.xlsx")
ws = wb['Sheet1']

score_rate = {'C':0.3, 'D':0.35, 'E':0.34, 'F':1}

i=2
while True:
    total=0
    if ws['A'+str(i)].value == None:
        break

    for k,v in score_rate.items():
        total+=ws[k+str(i)].value *v

    ws['G'+str(i)].value = total

    score.append((ws['G'+str(i)].value))
    i+=1

stu_nums=len(score)

ratio = {'A':0.3, 'B':0.7, 'C':1}
stu_grade = { i+2:[s,stu_nums,''] for i,s in enumerate(score) }

score.sort()

total =0
for k,v in ratio.items():

    cnt=0
    for i, value in stu_grade.items():
        s = value[0]
        rank = stu_nums-bisect_right(score,s)+score.count(s)
        stu_grade[i][1] = rank

        if rank<=int(stu_nums*v) and stu_grade[i][2]=="":
            stu_grade[i][2] = k+"0"
            cnt+=1

    for i in stu_grade.keys():
        rank = stu_grade[i][1]
        grade = stu_grade[i][2]

        if rank<=int(total+cnt*0.5) and grade == k+"0":
            stu_grade[i][2]= k+"+"

    total+=cnt

for i in stu_grade.keys():
    ws['H'+str(i)] = stu_grade[i][2]
wb.save(filename="student.xlsx")
wb.close()

